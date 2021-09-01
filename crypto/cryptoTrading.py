"""
    df = create_dataFrame(data_)
    Fagner Nunes 07/2021
    trading bot using binance API
"""
import sys
import hmac
from urllib.parse import urlencode, urljoin
import hashlib
import numpy as np
import pandas as pd 
from pandas import ExcelWriter
import requests as r  
import os
import config
from binance.spot  import Spot
import logging
from binance.lib.utils import config_logging
import matplotlib.pyplot as plt
import matplotlib as mpl
import datetime
import queue
import threading
import time
import logging
import websocket
import asyncio
import json
from binance.websocket.spot.websocket_client import SpotWebsocketClient as Client
from matplotlib.animation import FuncAnimation
import concurrent.futures
import multiprocessing as mp
from multiprocessing import freeze_support,Manager
from binance.error import ClientError
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

"""
    GETTING Arguments 
"""
args = sys.argv

#import testnet keys
testkey = config.TEST_KEY
testSecret = config.TEST_SECRET_KEY

# import keys
BINANCE_KEY= config.BINANCE_KEY
BINANCE_SECRET_KEY = config.BINANCE_SECRET_KEY

test_url = 'https://testnet.binance.vision'
#creates client obj
client = Spot(key=BINANCE_KEY,secret = BINANCE_SECRET_KEY)
#client = Spot(testkey, testSecret,base_url=test_url)

headers = {
    'X-MBX-APIKEY': testkey
}
plt.style.use('fivethirtyeight')

   #colum in dataframe
my_columns = [

            "open-time",
            "open",
            "high",
            "low",
            "close",
            "EMA9",
            "EMA6",
            "EMA3",
            "WMA",
            "BUY",
            "SELL",
            "ATR",
            "portfolio"
        ]

def place_order(symbol,side,typee,quantity):
    symbol = symbol.upper()
    print(side)
    params = {    
    "symbol": symbol,
    "side": side,
    "type": typee,
    "quantity": quantity,
    }
    response = client.new_order(**params)
    return response

def cancel_all_open_orders(symbol):
    orders = client.get_open_orders(symbol)
    for order in orders:
        response = client.cancel_order(symbol, orderId=order['orderId'])
        print("order canceled")

NaN = np.nan
# fuction to get market change in last 24h
def change24H():
    my_columns =  ["ticker","percentile",'volume']
    #url = 'https://api.binance.com/api/v1/ticker/24hr?'
    url = 'https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=100&page=1&sparkline=false&price_change_percentage=1h'
    response =  r.get(url).json()
    market = pd.DataFrame(columns = my_columns)

    for data in response:
        market = market.append(
            pd.Series([
                data["symbol"],
                data["price_change_percentage_1h_in_currency"],
                data["total_volume"]

                ],
                    index=my_columns
                ),
            ignore_index = True

            )

    market.sort_values("percentile",ascending= False,inplace= True)
    symbol = market.iloc[0]
    
    return symbol['ticker']+"usdt"





# gets data from binance
def get_pastData(symbol,interval):
    symbol =symbol.upper()
    data = r.get(f"https://api.binance.com/api/v3/klines?symbol={symbol}&interval={interval}&limit=350").json()
    return data 


def create_dataFrame(data_set):
 
    # creates data frame
    graph =  pd.DataFrame(columns = my_columns)


    # creates a NaN to avoid errors  
    NaN = np.nan
    # populates dataframe
    for candle in data_set:
    
        graph = graph.append(
            pd.Series([
                datetime.datetime.utcfromtimestamp(float(candle[0])/1000),
                candle[1],
                float(candle[2]),
                float(candle[3]),
                float(candle[4]),
                'N/A',
                'N/A',
                'N/A',
                'N/A',
                NaN,
               NaN,
               NaN,
               NaN
            
                ],
                    index=my_columns
                ),
            ignore_index = True

            )
    return graph
# this creates an array with integers 1 to 10 included    
weights = np.arange(1,11)



# calculating the wma 10
"""

A lambda function is a small anonymous function. 
A lambda function can take any number of arguments, but can only have one expression.
lambda arguments : expression

.apply() lets us create and pass any custum function to a ROLLING window

.dot product of two arrays.
"""

def calcWMA(n):
    wma10 = graph['close'].rolling(n).apply(lambda prices: np.dot(prices, weights)/weights.sum(),raw = True)

    graph['WMA'] = np.round(wma10, decimals = 3)


# calculating EMA
"""
A 10 day ema has a smoothingFactor/alfa = 2/(10+1) whichs is aprox 0.1818
"""

def calc_ema(dataset,n):
    ema = dataset['close'].ewm(span=n).mean()
    return ema

#ema9 = graph['close'].ewm(span=9).mean()
#ema6 = graph['close'].ewm(span=6).mean()
#ema3 = graph['close'].ewm(span=3).mean()

def calculate_atr(DATA,n):
    high_low = DATA['high'] - DATA['low']
    high_cp= np.abs(DATA['high'] - DATA['close'].shift())
    low_cp = np.abs(DATA['low'] - DATA['close'].shift())
    dp = pd.concat([high_low,high_cp,low_cp],axis=1)
    true_range = np.max(dp,axis=1)
    atr = true_range.rolling(n).mean()
    return atr


# function that prints change in portfolio
def getPortfolio(v1,v2):
    percentile = ((v2-v1)/v1)*100
    print(f"profit: {percentile}%")

""" ATR average of the last 14 candles (1.5 to 1)"""

# algorithm to decide whent to buy and when to sell shares
def buy_sell(graph,port):
  
    # global variables 
    position = False    
    stop_loss = 0
    profit = 0
    portfolio = port
    starting_portfolio = portfolio
    shares = 0

    
    
    for row in graph.index[15:]:
        
        
        EMA3 = graph.loc[row,'EMA3']
        EMA6 = graph.loc[row,'EMA6']
        EMA9 = graph.loc[row,'EMA9']
        price = graph.loc[row,'close']
        time =  graph.loc[row, 'open-time']
        
        if EMA3 > EMA6 and EMA3 > EMA9 and position==False:
        
            stop_loss = price - (graph.loc[row,'ATR']*1.5)
            profit = price + (graph.loc[row,'ATR']*3)  
            shares  = portfolio/price
            portfolio = 0
            graph.loc[row,"BUY"] = price
            position = True 
        
        if  price>=profit and position==True or price <= stop_loss and position==True:
            graph.loc[row,"SELL"] = price
            portfolio = shares*price
            position = False
            graph.loc[row,"portfolio"] = portfolio

            
            #getPortfolio(starting_portfolio,portfolio)               
        
    return graph

#getPortfolio(starting_portfolio,portfolio)

#with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
#    print(graph[['open-time','ATR']])

# function creating console
def console(q):
    
    # infinit loop
    while 1:
        # get input
        cmd = input("> ")
        q.put(cmd)

        #leave infinit loop
        if cmd == "quit":
            break
#actios
def simulation():
    #getting symbol
    symbol = input("enter symbol: ")
    interval = input("enter symbol: ")

    #testing functions
    pastDATA = get_pastData(symbol,interval)
    graph = create_dataFrame(pastDATA,symbol)
    ema9 = calc_ema(graph,9)
    ema6 = calc_ema(graph,6)
    ema3 = calc_ema(graph,3)

    atr = calculate_atr(graph,14) 
    graph['ATR'] = atr
    #adding ema to dataframe
    graph['EMA9'] = np.round(ema9, decimals = 3)
    graph['EMA6'] = np.round(ema6, decimals = 3)
    graph['EMA3'] = np.round(ema3, decimals = 3)

    graph = buy_sell(graph,100)

    #ploting graph
    plot1 =plt.figure(1)
    plt.plot(graph['open-time'],graph['close'],label = "Price",zorder=1)
    plt.plot(graph['open-time'],graph['EMA9'], label= "ema9",zorder=1)
    plt.plot(graph['open-time'],graph['EMA6'], label= "ema6", zorder=1)
    plt.plot(graph['open-time'],graph['EMA3'], label= "ema3",zorder=1)
    
    plot2= plt.figure(2)
    plt.plot(graph['open-time'],graph['close'],label = "Price",zorder=1)
    plt.scatter(graph['open-time'],graph['BUY'],label="buy")
    plt.scatter(graph['open-time'],graph['SELL'],label="sell")        

    plot3 = plt.figure(3)
    plt.plot(graph["open-time"], graph['portfolio'], marker = 'o',label = "Portfolio")



    plt.xlabel("time")
    plt.ylabel("price")

    plt.legend()
    plt.show()

#every time websocet recieves a msg this function is called
# and List is updated so graph can update 
def on_message(List,sale,starting_data,symbol,interval,ws,message):
    
    message2 = json.loads(message) 
    sell_action,take_profit,stop_loss = sell(float(message2['k']['c']),starting_data,symbol,interval) #calls sell func if conditons are met
        
    if sell_action: #if sell returns true sale data list updated
        data = message2['k']
        NaN = np.nan
        time = datetime.datetime.utcfromtimestamp(message2['E']/1000)
        sale.append([time,float(data['c']),float(starting_data['portfolio']),take_profit,stop_loss])
        if interval == "1m":
            starting_data['skip']=3
            print(starting_data['skip'])
                
    if message2['k']['x']==True: #every time a candel closes main data list is updated
        data = message2['k']
        NaN = np.nan
        time = datetime.datetime.utcfromtimestamp(message2['E']/1000)    
        List.append([time,float(data['o']),float(data['h']),float(data['l']),float(data['c']),NaN,NaN,NaN,NaN,NaN,NaN,NaN,NaN])
        print('1 minute')        
    
def on_close():
    print('FAGNER, CONNECTION HAS BEEN CLOSED!!!')
           
#initializes websocket args: symbol and List (List shared between process )           
def getCandels(symbol,interval,List,sale,starting_data):
    
    url = f'wss://stream.binance.com:9443/ws/{symbol}@kline_{interval}'
    ws= websocket.WebSocketApp(url, on_message=on_message,on_close=on_close) # starts websockets
    ws.on_message = lambda *x: on_message(List,sale,starting_data,symbol,interval, *x) # passes arguments to func that is called every time we receive data from sockets
    ws.run_forever()


def buy(ema3,ema6,ema9,price,time,starting_data,atr,symbol,interval):
    
    if ema3 > ema6 and ema3 >ema9 and starting_data['position']==False and starting_data==0:
    #if starting_data['position']==False: # checks if there is active position
                
        shares = starting_data['portfolio']/price
        shares = round(shares,starting_data['step_size'])
        response =place_order(symbol,'BUY','MARKET',shares) #calls funct that places order   
        
        if  response['status'] != 'filed': #if order is placed properly 
            
            #updates starting data
            payed_price = sum(float(fill['price'])*float(fill['qty']) for fill in response['fills'])
            starting_data['stop_loss'] = price - ( atr * 1.5 )
            starting_data['profit'] = price + (atr * 3)
            #starting_data['stop_loss'] = price - 1
            #starting_data['profit'] = price + 1
            starting_data['shares'] = float(response['executedQty'])
            starting_data['portfolio'] = 0
            starting_data['position'] = True
            price_share= price*starting_data['shares']
            print(f'price: {price_share}  -- payed_price {payed_price}')
            return True # returns that it was placed
        
        else:
            
            print('order not filled')    
            return False  # returns that order wasnt placed

    return False    #if everything goes wrong, returs that order wasnt placed  

def sell(price, starting_data,symbol,interval):
    
    profit = starting_data['profit']
    position = starting_data['position'] 
    stop_loss = starting_data['stop_loss']
    shares = starting_data['shares']
    shares = round(shares,starting_data['step_size'])
    is_profit=0
    is_loss =0
    
    # checks that there is a position and conditions for sale a met
    if price >= profit and position==True or price <= stop_loss  and position == True:
        
        if price>= profit:
            is_profit=1
            is_loss=0
        else:
            is_profit=0
            is_loss=1
        response = place_order(symbol,'SELL','MARKET',shares) # calls order func
        
        if response: #if order is placed  starting data list is updated

            portfolio = starting_data['shares']*price
            starting_data['portfolio']=portfolio
            starting_data['position'] = False # sets postion to no longer active
                       
            return  True, is_profit, is_loss; # returns that order was completed 
    
    return False ,is_profit,is_loss;     # returns that order failed   

#Animate function is called by ploting function every second
def animate(self,List,sale,df,sale_df,starting_data,symbol,interval,df_csv_name):
    
    data =list(List) #turns proxy main list into local list
    sale_data = list(sale) #turns proxy main list into local list
    sale_df_size=len(sale_df.index) #gests sales  dataframe size 
    sale_data_size = len(sale_data) #get sale list size
    size = len(data) # gets main  data list size .. t
    df_size = len(df.index) # get main dataframe size

    if size> 0: # if main data size > 0

        if  df.loc[df_size-1,'open-time'] != data[size-1][0]: #checks if there is new row in data list/ checks for new candel
            
            df.loc[df_size]= data[size-1] # adds new row to main dataframe
            # Calculates new EMA
            ema9 = calc_ema(df,9)
            ema6 = calc_ema(df,6)
            ema3 = calc_ema(df,3)
            atr = calculate_atr(df,14) #Calculates new ATR
            df['ATR'] = atr
            df['EMA9'] = ema9
            df['EMA6'] = ema6
            df['EMA3'] = ema3
            
            if not starting_data['position'] and  pd.isna(df.loc[df_size,'SELL']): # if there is no active positon
                # runs buy func returns true or false
                if interval == '1m':
                    starting_data['skip']=starting_data['skip'] - 1
                
                buy_action = buy(df.loc[df_size,'EMA3'],df.loc[df_size,'EMA6'],df.loc[df_size,'EMA9'],df.loc[df_size,'close'],df.loc[df_size,'open-time'],starting_data,df.loc[df_size,'ATR'],symbol,interval)
            
                if buy_action:  
                    df.loc[df_size,'BUY'] = df.loc[df_size,'close']#adds buy price to data frame
            
            wb = openpyxl.load_workbook(df_csv_name)
            ws = wb.active
            startRow = ws.max_row +1
            newRow =df.iloc[-1].tolist()
            ws.append(newRow)
            wb.save(filename=df_csv_name)
            wb.close()
# if new sale row, row is added to sale df.. sale has its own df, so it can be ploted properly with price plot
    if sale_data_size>0 and sale_df_size ==0 or sale_data_size>0 and sale_df.loc[sale_df_size-1,'time'] < sale[sale_data_size-1][0]:     
        sale_df.loc[sale_df_size] = sale_data[sale_data_size-1] 
        wb = openpyxl.load_workbook(df_csv_name)
        ws = wb.get_sheet_by_name('portfolio')
        startRow = ws.max_row +1
        newRow = sale_data[sale_data_size-1]
        ws.append(newRow)
        wb.save(filename=df_csv_name)
        wb.close()        
        print(sale_df)
        
    # data potting 
   
    plt.subplot(2,1,1) 
    plt.cla()
    plt.plot(df['open-time'],df['close'],label = "Price")
    plt.scatter(df['open-time'],df['BUY'],label = "Buy",color='green')
    plt.scatter(sale_df['time'],sale_df['price'],label='price', color='red')
    
    plt.subplot(2,1,2)
    plt.cla()
    plt.scatter(sale_df['time'],sale_df['portfolio'],label ='portfolio',color='black')
    plt.legend(loc='upper left')
    plt.tight_layout()


"""

fucion that plots the graph with previous data and initializes animate function.
function meant to be used with Multporcessing.

"""
def ploting(List,sale,starting_data,symbol, interval):
    
    data_set= get_pastData(symbol,interval) #gest past data so it can be ploted
    df = create_dataFrame(data_set) # creates data frame with past data
    sale_df = pd.DataFrame(columns = ["time","price",'portfolio','take-profit','stop-loss'])#creates data frame to host sales data
    ema9 = calc_ema(df,9)
    ema6 = calc_ema(df,6)
    ema3 = calc_ema(df,3)
    
    df['EMA9'] = np.round(ema9, decimals = 3)
    df['EMA6'] = np.round(ema6, decimals = 3)
    df['EMA3'] = np.round(ema3, decimals = 3)    
    atr = calculate_atr(df,14)
    df['ATR'] = atr

    now = datetime.datetime.now()
    current_time = now.strftime("%d%b%Y_%H%M")
    
    df_csv_name = f'{current_time}_df.xlsx'
    sale_csv_name = f'{current_time}_sale.csv'
    directory = os.getcwd()
    directory = f'{directory}/trix/{df_csv_name}'
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r) 
    wb.create_sheet('portfolio')
    ws2 = wb.get_sheet_by_name('portfolio')

    for r in dataframe_to_rows(sale_df, index=False, header = True):
        ws2.append(r)
    wb.save(directory)
    ani = FuncAnimation(plt.gcf(), animate,fargs=(List,sale,df,sale_df,starting_data,symbol,interval,directory), interval=1000)# animates plot everysecond by caling animate func
    plt.show()

def run_sockets(List,sale,starting_data,symbol,interval): #process
    getCandels(symbol,interval,List,sale,starting_data)
   
   
def foo():
    print("TESTINGG")

def invalid_input():
    print("invalid command :( ")

# main
def main():
    #actions connected to function
    cmd_actions = {'foo':foo,'simulation':simulation,"tradingTest":trading_test}
    
    cmd_queue =  queue.Queue()

    dj = threading.Thread(target = console, args = {cmd_queue})
    dj.start()

    while 1:
        ccmd = cmd_queue.get()
        if cmd == 'quit':
            break
        action = cmd_actions.get(cmd, invalid_input)
        action()

def trade():
    
    symbol = input("enter symbol: ")
    interval = input("enter interval to trade in: ")
    manager = mp.Manager()
    symbol_info = r.get(f'https://api.binance.com/api/v3/exchangeInfo?symbol={symbol.upper()}').json()
    List = manager.list() #list to be shared between processes
    sale = manager.list()#list used to store a 
    starting_data = manager.dict() #dict that holds starting data 
    portfolio =input('enter value to be traded: ')
    for filt in symbol_info['symbols'][0]['filters']:
        if filt['filterType'] == 'LOT_SIZE':
            step_size = filt['stepSize'].find('1') - 2
            starting_data['step_size'] = step_size
            print(f'step_size: {step_size}')
            break
   
    
    # populating dict 
    starting_data['skip'] = 0
    starting_data['position'] = False
    starting_data['stop_loss'] = 0
    starting_data['profit'] =0
    starting_data['portfolio']= float(portfolio)
    starting_data['starting_portfolio'] = float(portfolio)
    starting_data['shares'] = 0
    freeze_support() # seems to be mandatory
    p1=mp.Process(target=run_sockets, args=(List,sale,starting_data,symbol,interval)) #creates process 1
    p2=mp.Process(target=ploting, args=(List,sale,starting_data,symbol,interval))   # creates process 2
    # starts process 
    p1.start()
    p2.start()
    p1.join()
    p2.join()
    
def menu(args):
    if args[1] == '-h':
        print()
        print("""
               1- cancel all open orders for specific symbol:
                  python3 cryptoTrading.py cancel_all <SYMBOL>
               
               2- get account details:
                  python3 cryptoTrading.py account                  
                
               3- sells everthing:
                  python3 cryptoTrading.py sell_all <SYMBOL>  
                  """)
    if args[1] == 'cancel_all':
        if len(args) !=3:
            print('python3 cryptoTrading.py cancel_all <SYMBOL>')
        else:
            cancel_all_open_orders(args[2].upper())
    if args[1]=='account':
        if len(args) !=2:
            print('python3 cryptoTrading.py account')
        else:
            print(client.account())
    if args[1]=='trade':
        if len(args) !=2:
            print('python3 cryptoTrading.py simulate')
        else:
            trade()

    if args[1]=='sell':
        if len(args)!=2:
            print('python3 cryptoTrading.py sell')
        else:
            print(client.account())
            symbol = input('enter symbol: ')
            qty = input('enter qty: ')
            response = place_order(symbol,'SELL','MARKET',qty)
            print(response)
if __name__ == '__main__':
    if len(args)>1:        
        menu(args)
        

